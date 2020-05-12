import openpyxl
from openpyxl import Workbook
import os
from datetime import date

def write_to_column(column_num, data):
	for column in sheet_2.iter_cols(min_col=column_num, max_col=column_num):
		for each in column:
			if each.value == None:
				each.value = data
				work_book_2.save('TicketData.xlsx')
				return
	

		

work_book = openpyxl.load_workbook('TicketExport.xlsx')
sheet = work_book.get_sheet_by_name('Sheet1')
work_book_2 = Workbook()
sheet_2 = work_book_2.active
work_book_3 = openpyxl.load_workbook('PropHistory.xlsx')
sheet_3 = work_book_3.get_sheet_by_name('Sheet1')
work_book_4 = openpyxl.load_workbook('TicketTypeHistory.xlsx')
sheet_4 = work_book_4.get_sheet_by_name('Sheet1')
prop_data_dict = {}
ticket_type = {}
today = date.today()
today = f"{today.month}-{today.day}-{today.year}"
total_tickets = 0


try:
	os.remove('CurrentlyOpen.txt')
except:
	pass
sheet_3.cell(column=sheet_3.max_column+1, row=1, value=today)
work_book_3.save('PropHistory.xlsx')
sheet_4.cell(column=sheet_4.max_column+1, row=1, value=today)
work_book_4.save('TicketTypeHistory.xlsx')

for column in sheet.iter_cols(min_col=1, max_col=1):
	for each in column:
		sheet_2.append([None])
		each = each.value
		if each[0].isdigit():
			data = each[0:3]
			write_to_column(1, data)

for column in sheet.iter_cols(min_col=4, max_col=4):
	for each in column:
		sheet_2.append([None])
		each = each.value
		if each.startswith('Task') == False:
			data = each
			write_to_column(2, data)	

for row in sheet_2.iter_rows():
	if row[0].value == None:
		break
	if row[0].value in prop_data_dict:
		prop = prop_data_dict[row[0].value]
		prop.append(row[1].value)
	else:
		prop_data_dict[row[0].value] = [row[1].value]
	if row[1].value in ticket_type:
		ticket_type[row[1].value] = ticket_type[row[1].value]+1
	else:
		ticket_type[row[1].value] = 1


for key in prop_data_dict:
	row_num = 1
	total_tickets += len(prop_data_dict.get(key))
	for each in sheet_3.iter_rows():
		if str(key) in str(each[0].value):
			amount = len(prop_data_dict.get(key))
			sheet_3.cell(column=sheet_3.max_column, row=row_num, value=amount)
			work_book_3.save('PropHistory.xlsx')
			row_num = 0
			break
		else:
			row_num +=1

for key in ticket_type:
	row_num = 0
	for each in sheet_4.iter_rows():
		row_num +=1
		if str(key) == str(each[0].value):
			amount = ticket_type.get(key)
			sheet_4.cell(column=sheet_4.max_column, row=row_num, value=amount)
			work_book_4.save('TicketTypeHistory.xlsx')
			row_num = 0
			break
		elif str(each[0].value) == 'TOTAL':
			sheet_4.cell(column=sheet_4.max_column, row=row_num, value=total_tickets)
		
			




work_book.close()
work_book_2.close()
work_book_3.close()
work_book_4.close()


with open('CurrentlyOpen.txt', 'w+') as out_file:
	out_file.write(f"Total Tickets Today -- {total_tickets}\n")
	for each in sorted(prop_data_dict.keys()):
		out_file.write(f"{each} ({len(prop_data_dict.get(each))} tickets) : {str(sorted(prop_data_dict.get(each)))}\n")
os.remove('TicketData.xlsx')